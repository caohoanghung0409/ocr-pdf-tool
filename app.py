import streamlit as st
import pytesseract
from pdf2image import convert_from_bytes
import pandas as pd
import re
import tempfile
import os
import time
import base64
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

# =========================
# CONFIG
# =========================
st.set_page_config(page_title="THL PDF TO EXCEL", layout="wide")

# =========================
# SESSION
# =========================
if "processing" not in st.session_state:
    st.session_state.processing = False
if "done" not in st.session_state:
    st.session_state.done = False
if "clear_uploader" not in st.session_state:
    st.session_state.clear_uploader = False
if "last_uploaded_names" not in st.session_state:
    st.session_state.last_uploaded_names = []
if "excel_file" not in st.session_state:
    st.session_state.excel_file = None

# =========================
# STYLE (GIỮ NGUYÊN)
# =========================
st.markdown("""
<style>
header, #MainMenu, footer {visibility: hidden;}
.block-container {padding-top: 0.5rem !important;}
.stApp { background: #f1f5f9; }

.header {
    font-size:22px;
    font-weight:700;
    margin-bottom:10px;
}

[data-testid="stFileUploader"] {
    border: 2px dashed #93c5fd;
    padding: 25px;
    border-radius: 18px;
    background: white;
    transition: 0.3s;
}
[data-testid="stFileUploader"]:hover {
    border-color:#3b82f6;
}

div.stButton > button {
    background: linear-gradient(135deg,#3b82f6,#22c55e);
    color:white;
    border:none;
    border-radius:12px;
    padding:12px 24px;
    font-weight:600;
    font-size:15px;
    box-shadow:0 4px 14px rgba(0,0,0,0.15);
    transition: all 0.25s ease;
}
div.stButton > button:hover {
    transform: translateY(-2px) scale(1.02);
}

.new-btn button {
    background: linear-gradient(135deg,#f59e0b,#ef4444) !important;
}

.process-btn {
    margin-top: 25px;
    margin-bottom: 15px;
}

.file-row {
    margin-top:12px;
    padding:10px;
    border-radius:12px;
    background:white;
    box-shadow:0 2px 8px rgba(0,0,0,0.05);
}

.progress {
    height:8px;
    background:#e5e7eb;
    border-radius:999px;
    overflow:hidden;
    margin-top:6px;
}
.progress-bar {
    height:100%;
    background:linear-gradient(90deg,#3b82f6,#22c55e);
    transition: width 0.3s ease;
}

.global-wrap { margin:15px 0; }

.global-bar {
    position:relative;
    height:20px;
    background:#e5e7eb;
    border-radius:999px;
    overflow:hidden;
}

.global-fill {
    height:100%;
    border-radius:999px;
    transition: width 0.4s ease;
}

.global-fill::before {
    content:"";
    position:absolute;
    width:100%;
    height:100%;
    background: repeating-linear-gradient(
        45deg,
        rgba(255,255,255,0.2) 0,
        rgba(255,255,255,0.2) 10px,
        transparent 10px,
        transparent 20px
    );
    animation: move 1s linear infinite;
}

@keyframes move {
    from { background-position: 0 0; }
    to { background-position: 40px 0; }
}

.global-text {
    position:absolute;
    width:100%;
    text-align:center;
    font-size:12px;
    font-weight:700;
    top:0;
    line-height:20px;
}

.global-meta {
    display:flex;
    justify-content:space-between;
    font-size:13px;
    margin-bottom:6px;
}

.loading {
    font-size:14px;
    color:#475569;
    margin-top:10px;
}
</style>
""", unsafe_allow_html=True)

# =========================
# HEADER
# =========================
st.markdown('<div class="header">🚀 THL PDF → EXCEL </div>', unsafe_allow_html=True)

# =========================
# UPLOADER
# =========================
uploader_key = "uploader_1" if not st.session_state.clear_uploader else "uploader_2"

uploaded_files = st.file_uploader(
    "📂 Chọn file PDF",
    type=["pdf"],
    accept_multiple_files=True,
    key=uploader_key
)

current_names = [f.name for f in uploaded_files] if uploaded_files else []

if current_names != st.session_state.last_uploaded_names:
    st.session_state.processing = False
    st.session_state.done = False
    st.session_state.last_uploaded_names = current_names

# =========================
# OCR
# =========================
def ocr_extract(img):

    def read(image):
        text = pytesseract.image_to_string(image, lang='eng', config='--oem 3 --psm 6')
        sm = re.search(r"(SM\d{4}\.\d{4})", text)
        date = re.search(r"(\d{2}/\d{2}/\d{4})", text)
        return sm, date

    w, h = img.size

    for variant in [
        img,
        img.crop((0,0,w,int(h*0.4))),
        img.rotate(180, expand=True),
        img.rotate(180, expand=True).crop((0,0,w,int(h*0.4))),
        img.rotate(90, expand=True),
        img.rotate(270, expand=True)
    ]:
        sm, date = read(variant)
        if sm and date:
            return sm.group(1), date.group(1)

    return None, None

# =========================
# GLOBAL BAR (CHỈ HIỂN THỊ ETA)
# =========================
def render_global_bar(percent, speed, eta):

    eta_text = "Sắp xong..." if eta == 0 else f"{eta//60}m {eta%60}s"

    return f"""
<div class="global-wrap">
    <div class="global-meta">
        <div>⚡ {percent}%</div>
        <div>⏳ {eta_text}</div>
    </div>
    <div class="global-bar">
        <div class="global-fill" style="width:{percent}%; background:linear-gradient(90deg,#3b82f6,#22c55e);"></div>
        <div class="global-text">{percent}%</div>
    </div>
</div>
"""

# =========================
# PROCESS
# =========================
def extract_pdf(file, box, global_box, start_time, processed_pages, total_pages_all):

    results = []
    images = convert_from_bytes(file.read(), dpi=150)
    total_pages = len(images)

    for i, img in enumerate(images, start=1):

        processed_pages[0] += 1

        percent = int((i/total_pages)*100)
        global_percent = int((processed_pages[0] / total_pages_all) * 100)

        elapsed = time.time() - start_time
        speed = processed_pages[0] / elapsed if elapsed > 0 else 0
        remaining = total_pages_all - processed_pages[0]
        eta = int(remaining / speed) if speed > 0 else 0

        global_box.markdown(render_global_bar(global_percent, speed, eta), unsafe_allow_html=True)

        box.markdown(f"""
<div class="file-row">
📄 {file.name} — Trang {i}/{total_pages} ({percent}%)
<div class="progress">
<div class="progress-bar" style="width:{percent}%"></div>
</div>
</div>
""", unsafe_allow_html=True)

        sm, date = ocr_extract(img)

        if sm and date:
            results.append({
                "SM": sm,
                "Ngày": date,
                "Trang": i
            })

    return results

# =========================
# MAIN
# =========================
if uploaded_files:

    global_box = st.empty()
    boxes = [st.empty() for _ in uploaded_files]

    if not st.session_state.processing and not st.session_state.done:

        st.markdown('<div class="process-btn">', unsafe_allow_html=True)

        if st.button("🚀 Bắt đầu xử lý"):
            st.session_state.processing = True
            st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.processing:

        st.markdown('<div class="loading">⏳ Đang xử lý... vui lòng chờ</div>', unsafe_allow_html=True)

        start_time = time.time()

        total_pages_all = sum(len(convert_from_bytes(f.read(), dpi=50)) for f in uploaded_files)
        for f in uploaded_files:
            f.seek(0)

        processed_pages = [0]

        tmp_excel = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

        with pd.ExcelWriter(tmp_excel.name, engine='openpyxl') as writer:

            for i, f in enumerate(uploaded_files):

                data = extract_pdf(
                    f, boxes[i], global_box,
                    start_time, processed_pages, total_pages_all
                )

                if data:
                    df = pd.DataFrame(data)
                    df.insert(0, "STT", range(1, len(df)+1))

                    sheet_name = os.path.splitext(f.name)[0][:31]

                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(tmp_excel.name)

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 3

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

            for cell in ws[1]:
                cell.font = Font(bold=True)

        wb.save(tmp_excel.name)

        st.session_state.excel_file = tmp_excel.name
        st.session_state.processing = False
        st.session_state.done = True
        st.rerun()

# =========================
# DOWNLOAD
# =========================
if st.session_state.done:

    st.success("🎉 HOÀN THÀNH !!!")

    with open(st.session_state.excel_file, "rb") as f:
        data = f.read()

    b64 = base64.b64encode(data).decode()

    st.markdown(f"""
        <iframe src="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" style="display:none;"></iframe>
    """, unsafe_allow_html=True)

    st.markdown('<div class="new-btn">', unsafe_allow_html=True)
    if st.button("🔄 XỬ LÝ FILE MỚI"):
        st.session_state.done = False
        st.session_state.clear_uploader = not st.session_state.clear_uploader
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
